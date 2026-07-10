#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Import sheet_2.xlsx → debts.json cho agentThuno."""
from __future__ import annotations
import json
import re
import sys
from datetime import date, datetime
from pathlib import Path

import openpyxl

EXCEL_PATH = Path(r"C:\DuKickAgent\sheet_2.xlsx")
DEBTS_PATH = Path(__file__).parent.parent / "debt_data" / "debts.json"
TODAY = date(2026, 7, 3)
YEAR = 2026


def parse_date(s: str) -> date | None:
    """Parse DD/MM hoặc DD/MM/YYYY → date. Trả None nếu không hợp lệ."""
    if not s:
        return None
    s = s.strip()
    m = re.match(r"^(\d{1,2})/(\d{1,2})(?:/(\d{4}))?$", s)
    if not m:
        return None
    day, month = int(m.group(1)), int(m.group(2))
    year = int(m.group(3)) if m.group(3) else YEAR
    try:
        return date(year, month, day)
    except ValueError:
        return None


def extract_payment_dates(cell_texts: list[str]) -> list[tuple[date, str]]:
    """Quét list text cell, trả về [(due_date, ghi_chú)] các lần thanh toán."""
    results: list[tuple[date, str]] = []
    PAY_KW = r"(?:thanh\s*to[áa]n|t[aá]m\s*[ưu]́ng|thanh\s*to[áa]n\s*full|thanh\s*to[áa]n\s*đ[ợo]t)"
    DATE_PAT = r"(\d{1,2}/\d{1,2}(?:/\d{4})?)"

    for text in cell_texts:
        if not text:
            continue
        text_lower = text.lower()
        # Pattern: "thanh toán XX (DD/MM)" hoặc "DD/MM Thanh toán"
        for m in re.finditer(
            rf"{PAY_KW}[^(]*?\({DATE_PAT}\)", text, re.IGNORECASE | re.DOTALL
        ):
            d = parse_date(m.group(1))
            if d:
                results.append((d, text[:120].strip()))

        for m in re.finditer(
            rf"{DATE_PAT}\s+{PAY_KW}", text, re.IGNORECASE | re.DOTALL
        ):
            d = parse_date(m.group(1))
            if d:
                results.append((d, text[:120].strip()))

    # Dedup by date
    seen: set[date] = set()
    unique: list[tuple[date, str]] = []
    for item in sorted(results, key=lambda x: x[0]):
        if item[0] not in seen:
            seen.add(item[0])
            unique.append(item)
    return unique


def extract_amount(text: str) -> float:
    """Lấy số tiền từ text dạng '703,5M' '571M' '100tr' '500.000.000'."""
    # dạng XM (triệu → VND)
    m = re.search(r"([\d]+[,.]?\d*)\s*[Mm](?!\w)", text)
    if m:
        num_str = m.group(1).replace(",", ".")
        return float(num_str) * 1_000_000

    # dạng Xtr / triệu
    m = re.search(r"([\d]+[,.]?\d*)\s*(?:tr|triệu)", text, re.IGNORECASE)
    if m:
        num_str = m.group(1).replace(",", ".")
        return float(num_str) * 1_000_000

    # dạng số nguyên lớn có dấu chấm/phẩy phân cách
    m = re.search(r"\b(\d{1,3}(?:[.,]\d{3})+)\b", text)
    if m:
        num_str = re.sub(r"[.,]", "", m.group(1))
        return float(num_str)

    return 0.0


def extract_client(project_name: str) -> str:
    """Lấy tên khách hàng từ tên project (bỏ phần PM)."""
    # Bỏ phần "- Thái PM", "- Hoàng PM", "- Thuỷ PM", ...
    clean = re.sub(r"\s*-\s*\w+\s+PM\s*.*$", "", project_name, flags=re.IGNORECASE)
    # Bỏ phần "PENDING"
    clean = re.sub(r"\s*-?\s*PENDING\b.*", "", clean, flags=re.IGNORECASE)
    return clean.strip()


def week_start_date(week_range: str) -> date | None:
    """Lấy ngày đầu tuần từ chuỗi dạng '06/07-12/07' hay '29/06-05/07'."""
    m = re.match(r"(\d{1,2}/\d{1,2})", week_range)
    if m:
        return parse_date(m.group(1))
    return None


def build_debts() -> list[dict]:
    wb = openpyxl.load_workbook(str(EXCEL_PATH))
    ws = wb["2026"]

    # Lấy mapping col_index → week_range (hàng 2, từ col index 2 trở đi)
    week_headers: dict[int, str] = {}
    for idx, cell in enumerate(ws[2]):
        if cell.value and "/" in str(cell.value):
            week_headers[idx] = str(cell.value)

    debts: list[dict] = []
    uid = 1

    for row in ws.iter_rows(values_only=True):
        row_num = row[0]
        proj_name = row[1]

        if not (row_num and isinstance(row_num, (int, float)) and proj_name and str(proj_name).strip()):
            continue

        proj_name = str(proj_name).strip()
        client = extract_client(proj_name)

        # Tách text cells và numeric cells
        text_cells = [str(row[i]) for i in range(2, len(row)) if row[i] and isinstance(row[i], str)]
        numeric_cells = [
            (i, float(row[i]))
            for i in range(2, len(row))
            if isinstance(row[i], (int, float)) and float(row[i]) > 1_000_000
        ]

        payment_dates = extract_payment_dates(text_cells)

        if payment_dates:
            # Section 1 style: due date từ text, amount từ project name hoặc numeric cell
            amount = extract_amount(proj_name)
            if amount == 0 and numeric_cells:
                amount = numeric_cells[0][1]  # lấy numeric cell đầu tiên

            for due_date, note in payment_dates:
                status = "overdue" if due_date < TODAY else "pending"
                debts.append({
                    "id": f"DEBT-{uid:04d}",
                    "client_name": client,
                    "project": proj_name,
                    "amount": amount,
                    "currency": "VND",
                    "invoice_date": "",
                    "due_date": due_date.strftime("%Y-%m-%d"),
                    "status": status,
                    "contact_email": "",
                    "contact_phone": "",
                    "notes": note[:200],
                    "reminder_count": 0,
                    "last_reminder": None
                })
                uid += 1

        elif numeric_cells:
            # Section 2 style: amount là số, due date từ week-column header
            for col_idx, amount in numeric_cells:
                week_str = week_headers.get(col_idx, "")
                due_date = week_start_date(week_str) if week_str else None
                status = "overdue" if (due_date and due_date < TODAY) else "pending"
                debts.append({
                    "id": f"DEBT-{uid:04d}",
                    "client_name": client,
                    "project": proj_name,
                    "amount": amount,
                    "currency": "VND",
                    "invoice_date": "",
                    "due_date": due_date.strftime("%Y-%m-%d") if due_date else "",
                    "status": status,
                    "contact_email": "",
                    "contact_phone": "",
                    "notes": f"PPW/dự phóng: {week_str}",
                    "reminder_count": 0,
                    "last_reminder": None
                })
                uid += 1

        else:
            # Không có ngày lẫn amount → entry pending không rõ ngày
            all_text = " ".join(text_cells)[:200]
            amount = extract_amount(proj_name)
            debts.append({
                "id": f"DEBT-{uid:04d}",
                "client_name": client,
                "project": proj_name,
                "amount": amount,
                "currency": "VND",
                "invoice_date": "",
                "due_date": "",
                "status": "pending",
                "contact_email": "",
                "contact_phone": "",
                "notes": f"Chua ro ngay thanh toan. {all_text[:150]}",
                "reminder_count": 0,
                "last_reminder": None
            })
            uid += 1

    return debts


def main() -> None:
    sys.stdout.reconfigure(encoding="utf-8")
    debts = build_debts()
    db = {
        "version": 1,
        "updated_at": TODAY.isoformat(),
        "debts": debts
    }
    DEBTS_PATH.write_text(json.dumps(db, ensure_ascii=False, indent=2), encoding="utf-8")

    overdue = sum(1 for d in debts if d["status"] == "overdue")
    pending = sum(1 for d in debts if d["status"] == "pending")
    total_amount = sum(d["amount"] for d in debts)
    print(f"[OK] Import xong: {len(debts)} khoan cong no")
    print(f"   Qua han: {overdue} | Cho thanh toan: {pending}")
    print(f"   Tong tien co amount: {total_amount:,.0f} VND")
    print(f"   Saved to: {DEBTS_PATH}")


if __name__ == "__main__":
    main()
